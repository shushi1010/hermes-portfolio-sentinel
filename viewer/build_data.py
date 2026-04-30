#!/usr/bin/env python3
"""
Build portfolio_data.json from holdings_record.xlsx.

Usage:
  python3 viewer/build_data.py
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = BASE_DIR / "holdings_record.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parent / "portfolio_data.json"

NUMERIC_FIELDS = {
    "shares",
    "cost_price",
    "current_price",
    "previous_close",
    "market_value",
    "cost_basis",
    "profit_loss",
    "profit_loss_pct",
    "daily_change",
    "daily_change_pct",
    "daily_profit_loss",
    "weight_pct",
}


def to_float(value: object) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def calculate_totals(positions: list[dict], snapshot_time: str) -> dict:
    total_value = sum(to_float(p.get("market_value")) for p in positions)
    total_cost = sum(to_float(p.get("cost_basis")) for p in positions)
    total_profit_loss = total_value - total_cost
    total_daily_profit_loss = sum(to_float(p.get("daily_profit_loss")) for p in positions)
    for p in positions:
        p["weight_pct"] = to_float(p.get("market_value")) / total_value * 100 if total_value else 0.0
    return {
        "snapshot_time": snapshot_time,
        "position_count": len(positions),
        "total_market_value": total_value,
        "total_cost_basis": total_cost,
        "total_profit_loss": total_profit_loss,
        "total_profit_loss_pct": total_profit_loss / total_cost * 100 if total_cost else 0.0,
        "total_daily_profit_loss": total_daily_profit_loss,
        "total_daily_profit_loss_pct": total_daily_profit_loss / (total_value - total_daily_profit_loss) * 100
        if total_value != total_daily_profit_loss
        else 0.0,
    }


def load_latest_positions() -> tuple[list[dict], dict]:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Missing {EXCEL_PATH}. Run python3 monitor.py first.")

    wb = load_workbook(EXCEL_PATH, read_only=True)
    if "Snapshots" not in wb.sheetnames:
        raise ValueError("Sheet 'Snapshots' not found")
    ws = wb["Snapshots"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Snapshots sheet has no header row")

    headers = [normalize(h) for h in header_row]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {}
        for idx, name in enumerate(headers):
            if not name:
                continue
            value = row[idx] if idx < len(row) else ""
            item[name] = to_float(value) if name in NUMERIC_FIELDS else normalize(value)
        if item.get("symbol"):
            rows.append(item)

    if not rows:
        return [], calculate_totals([], "")

    latest_time = max(normalize(r.get("snapshot_time")) for r in rows)
    positions = [r for r in rows if normalize(r.get("snapshot_time")) == latest_time]
    positions.sort(key=lambda x: to_float(x.get("market_value")), reverse=True)
    totals = calculate_totals(positions, latest_time)
    return positions, totals


def main() -> None:
    positions, totals = load_latest_positions()
    payload = {
        "updated_at": totals.get("snapshot_time", ""),
        "totals": totals,
        "positions": positions,
        "markets": sorted({p.get("market", "") for p in positions if p.get("market")}),
        "sectors": sorted({p.get("sector", "") for p in positions if p.get("sector")}),
    }
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] Wrote {len(positions)} positions to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
