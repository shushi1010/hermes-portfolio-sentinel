#!/usr/bin/env python3
"""
Portfolio monitoring entrypoint.

Daily flow:
1. read holdings.csv
2. fetch latest quotes
3. calculate portfolio P/L and risk hints
4. append a snapshot to holdings_record.xlsx
5. export viewer/portfolio_data.json
6. write a Feishu-ready Markdown report
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent
HOLDINGS_FILE = BASE_DIR / "holdings.csv"
EXCEL_FILE = BASE_DIR / "holdings_record.xlsx"
VIEWER_JSON = BASE_DIR / "viewer" / "portfolio_data.json"
FEISHU_MESSAGE_FILE = BASE_DIR / "feishu_msg.md"
SNAPSHOT_JSON = BASE_DIR / "portfolio_snapshot.json"

YAHOO_CHART_API = "https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
REQUEST_TIMEOUT = 20

HEADERS = [
    "snapshot_time",
    "symbol",
    "name",
    "market",
    "sector",
    "shares",
    "cost_price",
    "current_price",
    "previous_close",
    "currency",
    "market_value",
    "cost_basis",
    "profit_loss",
    "profit_loss_pct",
    "daily_change",
    "daily_change_pct",
    "daily_profit_loss",
    "weight_pct",
    "risk_level",
    "action_hint",
    "latest_news",
    "llm_summary_cn",
    "notes",
]


@dataclass
class Holding:
    symbol: str
    name: str
    market: str
    shares: float
    cost_price: float
    sector: str = ""
    alert_below: float | None = None
    alert_above: float | None = None
    notes: str = ""


def to_float(value: object, default: float = 0.0) -> float:
    if value is None:
        return default
    text = str(value).strip().replace(",", "")
    if not text:
        return default
    try:
        number = float(text)
    except ValueError:
        return default
    if math.isnan(number) or math.isinf(number):
        return default
    return number


def fmt_money(value: float, currency: str = "") -> str:
    prefix = f"{currency} " if currency else ""
    return f"{prefix}{value:,.2f}"


def fmt_pct(value: float) -> str:
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.2f}%"


def fmt_plain_pct(value: float) -> str:
    return f"{value:.2f}%"


def create_sample_holdings() -> None:
    if HOLDINGS_FILE.exists():
        return
    rows = [
        {
            "symbol": "AAPL",
            "name": "Apple",
            "market": "US",
            "shares": "10",
            "cost_price": "180",
            "sector": "Technology",
            "alert_below": "160",
            "alert_above": "220",
            "notes": "示例持仓，请替换为你自己的股票",
        },
        {
            "symbol": "NVDA",
            "name": "NVIDIA",
            "market": "US",
            "shares": "5",
            "cost_price": "650",
            "sector": "Semiconductors",
            "alert_below": "580",
            "alert_above": "900",
            "notes": "示例持仓，请替换为你自己的股票",
        },
    ]
    with HOLDINGS_FILE.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def load_holdings() -> list[Holding]:
    create_sample_holdings()
    holdings: list[Holding] = []
    with HOLDINGS_FILE.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        required = {"symbol", "shares", "cost_price"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"holdings.csv missing required columns: {sorted(missing)}")

        for row in reader:
            symbol = (row.get("symbol") or "").strip()
            shares = to_float(row.get("shares"))
            cost_price = to_float(row.get("cost_price"))
            if not symbol or shares <= 0:
                continue
            alert_below = to_float(row.get("alert_below"), default=float("nan"))
            alert_above = to_float(row.get("alert_above"), default=float("nan"))
            holdings.append(
                Holding(
                    symbol=symbol,
                    name=(row.get("name") or symbol).strip(),
                    market=(row.get("market") or "").strip(),
                    shares=shares,
                    cost_price=cost_price,
                    sector=(row.get("sector") or "").strip(),
                    alert_below=None if math.isnan(alert_below) else alert_below,
                    alert_above=None if math.isnan(alert_above) else alert_above,
                    notes=(row.get("notes") or "").strip(),
                )
            )
    if not holdings:
        raise ValueError("holdings.csv has no valid holdings. Fill symbol, shares and cost_price.")
    return holdings


def fetch_quote(symbol: str) -> dict:
    url = YAHOO_CHART_API.format(symbol=quote(symbol, safe=""))
    response = requests.get(
        url,
        params={"range": "5d", "interval": "1d"},
        timeout=REQUEST_TIMEOUT,
        headers={"User-Agent": "Mozilla/5.0 hermes-portfolio-monitor"},
    )
    response.raise_for_status()
    payload = response.json()
    result = payload.get("chart", {}).get("result") or []
    if not result:
        error = payload.get("chart", {}).get("error") or {}
        message = error.get("description") or "empty quote response"
        raise RuntimeError(message)

    item = result[0]
    meta = item.get("meta") or {}
    indicators = item.get("indicators", {}).get("quote") or [{}]
    closes = [
        to_float(x, default=float("nan"))
        for x in indicators[0].get("close", [])
        if x is not None
    ]
    closes = [x for x in closes if not math.isnan(x)]

    current = to_float(meta.get("regularMarketPrice"), default=float("nan"))
    if math.isnan(current) and closes:
        current = closes[-1]

    previous = to_float(meta.get("chartPreviousClose"), default=float("nan"))
    if math.isnan(previous):
        previous = to_float(meta.get("previousClose"), default=float("nan"))
    if math.isnan(previous) and len(closes) >= 2:
        previous = closes[-2]
    if math.isnan(previous):
        previous = current

    if math.isnan(current) or current <= 0:
        raise RuntimeError("quote has no valid current price")

    return {
        "current_price": current,
        "previous_close": previous,
        "currency": meta.get("currency") or "",
        "exchange": meta.get("exchangeName") or meta.get("fullExchangeName") or "",
        "regular_market_time": meta.get("regularMarketTime") or "",
    }


def classify_risk(
    holding: Holding,
    profit_loss_pct: float,
    daily_change_pct: float,
    weight_pct: float,
    quote_error: str = "",
) -> tuple[str, str]:
    hints: list[str] = []
    risk_score = 0

    if quote_error:
        return "High", f"行情获取失败：{quote_error}"

    if weight_pct >= 40:
        risk_score += 2
        hints.append("单股仓位超过 40%，组合集中度较高")
    elif weight_pct >= 25:
        risk_score += 1
        hints.append("单股仓位超过 25%，需要关注集中度")

    if daily_change_pct <= -5:
        risk_score += 2
        hints.append("单日跌幅超过 5%，建议检查新闻和公告")
    elif daily_change_pct <= -3:
        risk_score += 1
        hints.append("单日跌幅超过 3%，短线波动加大")

    if profit_loss_pct <= -15:
        risk_score += 2
        hints.append("浮亏超过 15%，需要复盘持仓逻辑")
    elif profit_loss_pct <= -8:
        risk_score += 1
        hints.append("浮亏超过 8%，关注是否触及止损计划")

    if holding.alert_below is not None:
        hints.append(f"下方提醒价 {holding.alert_below:g}")
    if holding.alert_above is not None:
        hints.append(f"上方提醒价 {holding.alert_above:g}")

    if risk_score >= 3:
        risk = "High"
    elif risk_score >= 1:
        risk = "Medium"
    else:
        risk = "Low"

    if not hints:
        hints.append("价格和仓位处于常规观察区间")
    return risk, "；".join(hints)


def build_positions(holdings: list[Holding]) -> tuple[list[dict], dict]:
    snapshot_time = datetime.now().isoformat(timespec="seconds")
    positions: list[dict] = []

    for holding in holdings:
        quote_error = ""
        try:
            quote_data = fetch_quote(holding.symbol)
        except Exception as exc:
            quote_error = str(exc)
            quote_data = {
                "current_price": holding.cost_price,
                "previous_close": holding.cost_price,
                "currency": "",
                "exchange": "",
                "regular_market_time": "",
            }

        current_price = to_float(quote_data["current_price"])
        previous_close = to_float(quote_data["previous_close"], default=current_price)
        market_value = holding.shares * current_price
        cost_basis = holding.shares * holding.cost_price
        profit_loss = market_value - cost_basis
        profit_loss_pct = (profit_loss / cost_basis * 100) if cost_basis else 0.0
        daily_change = current_price - previous_close
        daily_change_pct = (daily_change / previous_close * 100) if previous_close else 0.0
        daily_profit_loss = daily_change * holding.shares

        positions.append(
            {
                "snapshot_time": snapshot_time,
                "symbol": holding.symbol,
                "name": holding.name,
                "market": holding.market or quote_data.get("exchange", ""),
                "sector": holding.sector,
                "shares": holding.shares,
                "cost_price": holding.cost_price,
                "current_price": current_price,
                "previous_close": previous_close,
                "currency": quote_data.get("currency", ""),
                "market_value": market_value,
                "cost_basis": cost_basis,
                "profit_loss": profit_loss,
                "profit_loss_pct": profit_loss_pct,
                "daily_change": daily_change,
                "daily_change_pct": daily_change_pct,
                "daily_profit_loss": daily_profit_loss,
                "weight_pct": 0.0,
                "risk_level": "",
                "action_hint": "",
                "latest_news": "",
                "llm_summary_cn": "",
                "notes": holding.notes,
                "quote_error": quote_error,
                "alert_below": holding.alert_below,
                "alert_above": holding.alert_above,
            }
        )

    total_value = sum(p["market_value"] for p in positions)
    total_cost = sum(p["cost_basis"] for p in positions)
    total_pl = total_value - total_cost
    total_day_pl = sum(p["daily_profit_loss"] for p in positions)

    holding_by_symbol = {h.symbol: h for h in holdings}
    for p in positions:
        p["weight_pct"] = (p["market_value"] / total_value * 100) if total_value else 0.0
        risk, hint = classify_risk(
            holding_by_symbol[p["symbol"]],
            p["profit_loss_pct"],
            p["daily_change_pct"],
            p["weight_pct"],
            p.get("quote_error", ""),
        )
        p["risk_level"] = risk
        p["action_hint"] = hint

    totals = {
        "snapshot_time": snapshot_time,
        "position_count": len(positions),
        "total_market_value": total_value,
        "total_cost_basis": total_cost,
        "total_profit_loss": total_pl,
        "total_profit_loss_pct": (total_pl / total_cost * 100) if total_cost else 0.0,
        "total_daily_profit_loss": total_day_pl,
        "total_daily_profit_loss_pct": (total_day_pl / (total_value - total_day_pl) * 100)
        if total_value != total_day_pl
        else 0.0,
    }
    return positions, totals


def load_or_create_workbook() -> openpyxl.Workbook:
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Snapshots" not in wb.sheetnames:
            wb.create_sheet("Snapshots")
        ws = wb["Snapshots"]
        if ws.max_row == 1 and not ws.cell(row=1, column=1).value:
            ws.append(HEADERS)
        return wb

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Snapshots"
    ws.append(HEADERS)
    header_fill = PatternFill(start_color="155E75", end_color="155E75", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [20, 12, 22, 12, 18, 10, 12, 13, 14, 10, 15, 15, 15, 14, 13, 14, 16, 12, 12, 36, 28, 36, 24]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    return wb


def append_snapshot_to_excel(positions: list[dict]) -> None:
    wb = load_or_create_workbook()
    ws = wb["Snapshots"]
    for p in positions:
        row = [p.get(h, "") for h in HEADERS]
        ws.append(row)
        for col in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(EXCEL_FILE)


def latest_positions_from_excel() -> tuple[list[dict], dict]:
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Missing {EXCEL_FILE}. Run python3 monitor.py first.")

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    ws = wb["Snapshots"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Snapshots sheet has no header")

    headers = [str(h) if h is not None else "" for h in header_row]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {name: row[idx] if idx < len(row) else "" for idx, name in enumerate(headers)}
        if item.get("symbol"):
            rows.append(item)
    if not rows:
        return [], {}

    latest_time = max(str(r.get("snapshot_time") or "") for r in rows)
    positions = [normalize_position(r) for r in rows if str(r.get("snapshot_time") or "") == latest_time]
    totals = calculate_totals(positions, latest_time)
    return positions, totals


def normalize_position(row: dict) -> dict:
    result = dict(row)
    numeric = [
        "shares",
        "cost_price",
        "current_price",
        "previous_close",
        "market_value",
        "cost_basis",
        "profit_loss",
        "profit_loss_pct",
        "daily_change",
        "daily_change_pct",
        "daily_profit_loss",
        "weight_pct",
    ]
    for key in numeric:
        result[key] = to_float(result.get(key))
    for key in ["symbol", "name", "market", "sector", "currency", "risk_level", "action_hint", "latest_news", "llm_summary_cn", "notes"]:
        result[key] = "" if result.get(key) is None else str(result.get(key))
    return result


def calculate_totals(positions: list[dict], snapshot_time: str) -> dict:
    total_value = sum(to_float(p.get("market_value")) for p in positions)
    total_cost = sum(to_float(p.get("cost_basis")) for p in positions)
    total_pl = total_value - total_cost
    total_day_pl = sum(to_float(p.get("daily_profit_loss")) for p in positions)
    for p in positions:
        p["weight_pct"] = (to_float(p.get("market_value")) / total_value * 100) if total_value else 0.0
    return {
        "snapshot_time": snapshot_time,
        "position_count": len(positions),
        "total_market_value": total_value,
        "total_cost_basis": total_cost,
        "total_profit_loss": total_pl,
        "total_profit_loss_pct": (total_pl / total_cost * 100) if total_cost else 0.0,
        "total_daily_profit_loss": total_day_pl,
        "total_daily_profit_loss_pct": (total_day_pl / (total_value - total_day_pl) * 100)
        if total_value != total_day_pl
        else 0.0,
    }


def export_viewer_json(positions: list[dict], totals: dict) -> None:
    positions = sorted(positions, key=lambda x: to_float(x.get("market_value")), reverse=True)
    payload = {
        "updated_at": totals.get("snapshot_time", ""),
        "totals": totals,
        "positions": positions,
        "markets": sorted({p.get("market", "") for p in positions if p.get("market")}),
        "sectors": sorted({p.get("sector", "") for p in positions if p.get("sector")}),
    }
    VIEWER_JSON.parent.mkdir(parents=True, exist_ok=True)
    VIEWER_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    SNAPSHOT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_feishu_message(positions: list[dict], totals: dict) -> str:
    leaders = sorted(positions, key=lambda p: to_float(p.get("daily_change_pct")), reverse=True)
    biggest_up = leaders[0] if leaders else None
    biggest_down = leaders[-1] if leaders else None
    high_risks = [p for p in positions if p.get("risk_level") == "High"]

    currency = positions[0].get("currency", "") if positions else ""
    lines = [
        f"📈 **持仓日报** | {datetime.now().date().isoformat()}",
        "",
        f"组合市值：**{fmt_money(totals.get('total_market_value', 0), currency)}**",
        f"今日盈亏：**{fmt_money(totals.get('total_daily_profit_loss', 0), currency)} ({fmt_pct(totals.get('total_daily_profit_loss_pct', 0))})**",
        f"累计盈亏：**{fmt_money(totals.get('total_profit_loss', 0), currency)} ({fmt_pct(totals.get('total_profit_loss_pct', 0))})**",
        "",
    ]

    if biggest_up:
        lines.append(
            f"涨幅最大：{biggest_up['symbol']} {fmt_pct(biggest_up['daily_change_pct'])}，"
            f"当前 {fmt_money(biggest_up['current_price'], biggest_up.get('currency', ''))}"
        )
    if biggest_down:
        lines.append(
            f"跌幅最大：{biggest_down['symbol']} {fmt_pct(biggest_down['daily_change_pct'])}，"
            f"当前 {fmt_money(biggest_down['current_price'], biggest_down.get('currency', ''))}"
        )

    lines.extend(["", "重点提醒："])
    if high_risks:
        for p in high_risks[:5]:
            lines.append(f"- {p['symbol']}：{p['action_hint']}")
    else:
        top_weight = sorted(positions, key=lambda p: to_float(p.get("weight_pct")), reverse=True)[:3]
        for p in top_weight:
            lines.append(f"- {p['symbol']} 仓位 {fmt_plain_pct(p['weight_pct'])}：{p['action_hint']}")

    lines.extend(["", "持仓明细："])
    for p in sorted(positions, key=lambda x: to_float(x.get("weight_pct")), reverse=True):
        lines.append(
            f"- {p['symbol']} {p.get('name', '')} | 仓位 {fmt_plain_pct(p['weight_pct'])} | "
            f"日涨跌 {fmt_pct(p['daily_change_pct'])} | 累计 {fmt_pct(p['profit_loss_pct'])}"
        )

    lines.append("")
    lines.append("数据已更新至 holdings_record.xlsx，网站数据已更新至 viewer/portfolio_data.json。")
    return "\n".join(lines)


def run_monitor() -> None:
    holdings = load_holdings()
    positions, totals = build_positions(holdings)
    append_snapshot_to_excel(positions)
    export_viewer_json(positions, totals)
    message = build_feishu_message(positions, totals)
    FEISHU_MESSAGE_FILE.write_text(message + "\n", encoding="utf-8")
    print(message)
    print(f"\n[OK] Excel: {EXCEL_FILE}")
    print(f"[OK] Viewer data: {VIEWER_JSON}")
    print(f"[OK] Feishu message: {FEISHU_MESSAGE_FILE}")


def rebuild_from_excel() -> None:
    positions, totals = latest_positions_from_excel()
    export_viewer_json(positions, totals)
    print(f"[OK] Rebuilt {VIEWER_JSON} from {EXCEL_FILE} ({len(positions)} positions)")


def main() -> None:
    parser = argparse.ArgumentParser(description="Monitor stock holdings and build portfolio reports")
    parser.add_argument("--build-viewer-data", action="store_true", help="rebuild viewer JSON from holdings_record.xlsx")
    parser.add_argument(
        "--sync-pending-state",
        action="store_true",
        help="compatibility no-op for old cron prompts; rebuilds viewer JSON",
    )
    args = parser.parse_args()

    try:
        if args.build_viewer_data or args.sync_pending_state:
            rebuild_from_excel()
        else:
            run_monitor()
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
