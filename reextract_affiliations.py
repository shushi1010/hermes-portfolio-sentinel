#!/usr/bin/env python3
"""
Re-extract affiliations from local PDFs and write them back to papers_record.xlsx.

Rules implemented for this batch:
- Read first 2 pages with pdfplumber
- Prefer footnote-style affiliation lines near the author block
- Repair CamelCase and glued words
- Merge hyphenated cross-line words
- Remove URLs/emails/body-text noise
- Output multiple affiliations separated by '; '
- Fallback to "未找到单位信息"
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

import pdfplumber


BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "papers_record.xlsx"
PAPERS_DIR = BASE_DIR / "papers"
TARGET_ROW_START = 52
TARGET_ROW_END = 201

ORG_KEYWORDS = {
    "university", "institute", "school", "college", "department", "laboratory",
    "lab", "centre", "center", "research", "academy", "hospital", "faculty",
    "polytechnic", "polytechnique", "technological", "technology", "technion",
    "google", "microsoft", "meta", "apple", "amazon", "ibm", "intel", "nvidia",
    "amd", "qualcomm", "samsung", "huawei", "tencent", "alibaba", "bytedance",
    "deepmind", "openai", "anthropic", "mistral", "cohere", "huggingface",
    "mit", "stanford", "harvard", "princeton", "yale", "berkeley", "cornell",
    "oxford", "cambridge", "eth", "epfl", "inria", "tum", "kaist", "postech",
    "cmu", "carnegie", "gatech", "purdue", "uiuc", "columbia", "caltech",
    "ucla", "ucsd", "toronto", "montreal", "tsinghua", "peking", "fudan",
    "zhejiang", "nanjing", "shanghai", "beihang", "sjtu", "ustc", "unist",
    "ntu", "nus", "renmin", "cas", "academy of military sciences", "istat",
}

AFFILIATION_HINTS = {
    "engineering", "science", "computer", "mathematics", "statistics", "ai",
    "artificial intelligence", "informatics", "information", "electrical",
    "electronic", "automation", "physics", "medicine", "medical", "business",
    "data", "robotics", "systems", "communication", "software",
}

SMALL_WORDS = [
    "of", "the", "and", "for", "in", "on", "at", "to", "by", "with", "from",
    "de", "du", "la", "le", "da", "del", "di",
]

NOISE_PATTERNS = [
    r"https?://\S+",
    r"www\.\S+",
    r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[A-Za-z]{2,}\b",
    r"\barxiv\b",
    r"\bcopyright\b",
    r"\bpreprint\b",
    r"\baccepted\b",
    r"\bpublication\b",
    r"\bfigure\b",
    r"\btable\b",
    r"\bappendix\b",
]

BAD_LINE_PATTERNS = [
    r"\babstract\b",
    r"\bintroduction\b",
    r"\bcontributions?\b",
    r"\brelated work\b",
    r"\bwe (?:study|show|propose|introduce|present|analyze|derive|establish)\b",
    r"\bour (?:method|analysis|results|experiments|framework)\b",
    r"\baccuracy\b",
    r"\bbenchmark\b",
    r"\bproof\b",
    r"\btheorem\b",
    r"\bresults?\b",
]

KNOWN_SUFFIXES = [
    "University", "Institute", "School", "College", "Department", "Laboratory",
    "Research", "Center", "Centre", "Hospital", "Sciences", "Technology",
    "Engineering", "Mathematics", "Physics", "Medicine", "Data", "Systems",
]


def normalize_text(text: str) -> str:
    text = text.replace("\u2013", "-").replace("\u2014", "-").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip(" ,;:-")
    return text


def fix_glued_words(text: str) -> str:
    text = normalize_text(text)
    text = re.sub(r"(?<=\d)(?=[A-Za-z])", " ", text)
    text = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", text)
    for word in SMALL_WORDS:
        text = re.sub(rf"(?i)([A-Za-z])({word})([A-Z])", r"\1 \2 \3", text)
        text = re.sub(rf"(?i)([A-Za-z])({word})([a-z])", r"\1 \2 \3", text)
    for suffix in KNOWN_SUFFIXES:
        text = re.sub(rf"(?<=[A-Za-z])(?={suffix}\b)", " ", text)
    text = re.sub(r"(?<=,)(?=\S)", " ", text)
    text = re.sub(r"(?<=;)(?=\S)", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" ,;:-")
    return text


def clean_candidate(text: str) -> str:
    text = normalize_text(text)
    text = re.sub(r"[\*\†\‡§¶‖#]+", " ", text)
    text = re.sub(r"(?<![A-Za-z])[\^]?\d+(?=[A-Za-z])", "", text)
    text = re.sub(r"^\s*[\^]?\d+[\)\].,:-]?\s*", "", text)
    for pattern in NOISE_PATTERNS:
        text = re.sub(pattern, " ", text, flags=re.IGNORECASE)
    text = fix_glued_words(text)
    text = re.sub(r"\([^)]*@[^\)]*\)", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" ,;:-")
    return text


def has_long_unspaced_token(text: str) -> bool:
    return bool(re.search(r"[A-Za-z]{18,}", text))


def is_noise_line(text: str) -> bool:
    low = text.lower()
    if not text:
        return True
    if any(re.search(p, low) for p in BAD_LINE_PATTERNS):
        return True
    if len(re.findall(r"[=<>±∑∫]", text)) > 0:
        return True
    if re.search(r"\b\d{4}\b", text) and not has_org_signal(text):
        return True
    return False


def has_org_signal(text: str) -> bool:
    low = text.lower()
    return any(k in low for k in ORG_KEYWORDS) or any(k in low for k in AFFILIATION_HINTS)


def looks_like_affiliation(text: str) -> bool:
    if not text or is_noise_line(text):
        return False
    if not has_org_signal(text):
        return False
    letters = len(re.findall(r"[A-Za-z]", text))
    if letters < 6:
        return False
    if sum(ch in ",;" for ch in text) >= 4 and "university" not in text.lower() and "institute" not in text.lower():
        return False
    return True


def split_compound_affiliations(text: str) -> list[str]:
    text = clean_candidate(text)
    if not text:
        return []
    pieces = [text]
    numbered = re.split(r"\s(?=\d+[\)\].,:-]?\s*[A-Z])", text)
    if len(numbered) > 1:
        pieces = numbered
    results: list[str] = []
    for piece in pieces:
        piece = clean_candidate(piece)
        if looks_like_affiliation(piece):
            results.append(piece)
    return results


def dedupe_affiliations(items: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        item = clean_candidate(item)
        if has_long_unspaced_token(item):
            item = fix_glued_words(item)
        item = normalize_text(item)
        if not looks_like_affiliation(item):
            continue
        key = re.sub(r"[^a-z0-9]+", "", item.lower())
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def words_to_lines(page) -> list[tuple[float, str]]:
    words = page.extract_words(
        x_tolerance=1.5,
        y_tolerance=3,
        keep_blank_chars=False,
        use_text_flow=True,
    )
    if not words:
        return []
    rows: list[list[dict]] = []
    for word in sorted(words, key=lambda w: (w["top"], w["x0"])):
        if rows and abs(rows[-1][0]["top"] - word["top"]) <= 2.5:
            rows[-1].append(word)
        else:
            rows.append([word])
    lines: list[tuple[float, str]] = []
    for row in rows:
        row = sorted(row, key=lambda w: w["x0"])
        parts = [w["text"] for w in row]
        text = " ".join(parts)
        text = normalize_text(text)
        if text:
            lines.append((row[0]["top"], text))
    return lines


def merge_hyphen_lines(lines: list[str]) -> list[str]:
    merged: list[str] = []
    i = 0
    while i < len(lines):
        line = lines[i].rstrip()
        while line.endswith("-") and i + 1 < len(lines):
            nxt = lines[i + 1].lstrip()
            line = line[:-1] + nxt
            i += 1
        merged.append(line)
        i += 1
    return merged


def extract_candidates_from_page1(pdf_path: Path) -> list[str]:
    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            return []
        page1 = pdf.pages[0]
        lines = words_to_lines(page1)
        if not lines:
            return []
        abstract_y = None
        for top, text in lines:
            if re.fullmatch(r"abstract", text, flags=re.IGNORECASE):
                abstract_y = top
                break
        if abstract_y is None:
            for top, text in lines:
                if re.match(r"^\d+(\.\d+)?\s+introduction\b", text, flags=re.IGNORECASE):
                    abstract_y = top
                    break
        top_lines = [text for top, text in lines if abstract_y is None or top < abstract_y]
        top_lines = merge_hyphen_lines(top_lines)

        candidates: list[str] = []
        for line in top_lines:
            cleaned = clean_candidate(line)
            if looks_like_affiliation(cleaned):
                candidates.extend(split_compound_affiliations(cleaned))
        return dedupe_affiliations(candidates)


def extract_candidates_from_first_two_pages(pdf_path: Path) -> list[str]:
    candidates: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages[:2]:
            lines = [text for _, text in words_to_lines(page)]
            lines = merge_hyphen_lines(lines)
            for line in lines:
                cleaned = clean_candidate(line)
                if looks_like_affiliation(cleaned):
                    candidates.extend(split_compound_affiliations(cleaned))
    return dedupe_affiliations(candidates)


def extract_affiliations(pdf_path: Path) -> str:
    if not pdf_path.exists():
        return "未找到单位信息"
    page1_affs = extract_candidates_from_page1(pdf_path)
    if page1_affs:
        return "; ".join(page1_affs)
    page2_fallback = extract_candidates_from_first_two_pages(pdf_path)
    if page2_fallback:
        return "; ".join(page2_fallback)
    return "未找到单位信息"


def main() -> None:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    if "pdf_local_path" not in idx and "pdf_filename" not in idx:
        raise RuntimeError("Excel missing pdf_local_path/pdf_filename column")

    for row in range(TARGET_ROW_START, TARGET_ROW_END + 1):
        arxiv_id = ws.cell(row, idx["arxiv_id"]).value
        if "pdf_local_path" in idx and ws.cell(row, idx["pdf_local_path"]).value:
            pdf_path = Path(ws.cell(row, idx["pdf_local_path"]).value)
        else:
            pdf_filename = ws.cell(row, idx["pdf_filename"]).value
            pdf_path = PAPERS_DIR / pdf_filename if pdf_filename else PAPERS_DIR / f"{arxiv_id}.pdf"
        affiliations = extract_affiliations(pdf_path)
        ws.cell(row, idx["affiliations"]).value = affiliations
        print(f"{row}\t{arxiv_id}\t{affiliations}")

    wb.save(EXCEL_PATH)


if __name__ == "__main__":
    main()
